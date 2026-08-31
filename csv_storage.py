import csv
import os
from datetime import datetime

def get_csv_path(season):
    """Get CSV file path for a season."""
    return os.path.join(os.path.dirname(__file__), f'SNIPES_{season}.csv')

def ensure_csv_exists(season):
    """Create CSV file if it doesn't exist."""
    csv_path = get_csv_path(season)
    if not os.path.exists(csv_path):
        with open(csv_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Sniper', 'Points', 'Snipee', 'Timestamp', 'Proof Link', 'Sniper ID', 'Snipee ID'])

def append_snipe(season, sniper, points, snipee, timestamp, proof_path, sniper_id, snipee_id):
    """Append a snipe to the CSV."""
    ensure_csv_exists(season)
    csv_path = get_csv_path(season)
    with open(csv_path, 'a', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([sniper, points, snipee, timestamp, proof_path, sniper_id, snipee_id])

def get_snipes(season):
    """Get all snipes from CSV (returns list of dicts)."""
    ensure_csv_exists(season)
    csv_path = get_csv_path(season)
    snipes = []
    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['Sniper']:  # Skip empty rows
                snipes.append(row)
    return snipes

def edit_snipe(season, sniper, snipee, points, field, new_value):
    """Edit a snipe in the CSV."""
    csv_path = get_csv_path(season)
    snipes = get_snipes(season)
    
    # Find and edit the snipe
    for snipe in snipes:
        if (snipe['Sniper'] == sniper and 
            snipe['Snipee'] == snipee and 
            snipe['Points'] == str(points)):
            snipe[field] = new_value
            break
    
    # Write back to CSV
    with open(csv_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['Sniper', 'Points', 'Snipee', 'Timestamp', 'Proof Link', 'Sniper ID', 'Snipee ID'])
        writer.writeheader()
        writer.writerows(snipes)

def delete_snipe(season, sniper, snipee, points):
    """Delete a snipe from the CSV."""
    csv_path = get_csv_path(season)
    snipes = get_snipes(season)
    
    # Remove the snipe
    snipes = [s for s in snipes if not (s['Sniper'] == sniper and s['Snipee'] == snipee and s['Points'] == str(points))]
    
    # Write back to CSV
    with open(csv_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['Sniper', 'Points', 'Snipee', 'Timestamp', 'Proof Link', 'Sniper ID', 'Snipee ID'])
        writer.writeheader()
        writer.writerows(snipes)

if __name__ == '__main__':
    # Test
    append_snipe('TEST2026', 'Alice', 2, 'Bob', '2026-08-31 12:00:00', '/path/to/proof.jpg', '123', '456')
    snipes = get_snipes('TEST2026')
    print(f"Snipes: {snipes}")

def merge_csv_to_excel(season, workbook):
    """Merge CSV snipes into Excel sheet for the current season only. Preserves all other sheets."""
    import openpyxl

    # Get or create the sheet for current season
    if season in workbook.sheetnames:
        sheet = workbook[season]
        # Delete all data rows (keep header in row 1)
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row)
    else:
        sheet = workbook.create_sheet(season)
        sheet.append(['Sniper', 'Points', 'Snipee', 'Timestamp', 'Proof Link', 'Sniper ID', 'Snipee ID'])

    # Read CSV and write to Excel (starts at row 2)
    snipes = get_snipes(season)
    for idx, snipe in enumerate(snipes, start=2):
        sheet.cell(row=idx, column=1).value = snipe['Sniper']
        sheet.cell(row=idx, column=2).value = int(snipe['Points'])
        sheet.cell(row=idx, column=3).value = snipe['Snipee']
        sheet.cell(row=idx, column=4).value = snipe['Timestamp']
        sheet.cell(row=idx, column=5).value = snipe['Proof Link']
        sheet.cell(row=idx, column=6).value = snipe['Sniper ID']
        sheet.cell(row=idx, column=7).value = snipe['Snipee ID']

    # All other sheets remain untouched
    return workbook
