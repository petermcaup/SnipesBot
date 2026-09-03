import discord
from discord import app_commands
from discord.ext import commands
import openpyxl
import json
from datetime import datetime
from private import private
import sys
import os
import asyncio
from concurrent.futures import ThreadPoolExecutor
import schedule
from google_drive_backup import upload_backup
from csv_storage import append_snipe, get_snipes, edit_snipe as csv_edit_snipe, delete_snipe as csv_delete_snipe, merge_csv_to_excel

# --- DYNAMIC PATHING ---
if getattr(sys, 'frozen', False):
    # Path of the .exe inside the 'dist' folder
    EXE_LOCATION = os.path.dirname(sys.executable)
    # Move UP one level to the main 'SnipesBot' folder
    BASE_DIR = os.path.dirname(EXE_LOCATION)
else:
    # If running as a .py script, assume it's already in the main folder
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- CONFIGURATION ---
TOKEN = private.token
OWNER_ID = int(private.owner_id) 

# These will now point to SnipesBot\SNIPESSTATS.xlsm instead of SnipesBot\dist\SNIPESSTATS.xlsm
EXCEL_FILE = os.path.join(BASE_DIR, 'SNIPESSTATS.xlsx')
REG_FILE = os.path.join(BASE_DIR, 'private', 'registrations.json')
PROOFS_DIR = os.path.join(BASE_DIR, 'proofs')
ROW_TRACKER_FILE = os.path.join(BASE_DIR, 'private', 'row_tracker.json')

# Ensure the private directory exists in the main folder
PRIVATE_DIR = os.path.join(BASE_DIR, 'private')
if not os.path.exists(PRIVATE_DIR):
    os.makedirs(PRIVATE_DIR)

# Ensure the proofs directory exists
if not os.path.exists(PROOFS_DIR):
    os.makedirs(PROOFS_DIR)

print(f"Bot starting... Working Directory: {BASE_DIR}")

# Thread pool for blocking I/O operations
executor = ThreadPoolExecutor(max_workers=1)

async def save_workbook_async(workbook):
    """Save workbook in a thread to avoid blocking the event loop."""
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(executor, workbook.save, EXCEL_FILE)

# --- DATA PERSISTENCE HELPERS ---

def load_data():
    """Loads season and registration data from JSON."""
    if not os.path.exists(REG_FILE):
        # Default state if no file exists
        return {"season": "FALL2026", "registrations": {}}
    with open(REG_FILE, 'r') as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {"season": "FALL2026", "registrations": {}}

def save_data(season, registrations):
    """Saves season and registration data to JSON."""
    data = {
        "season": season,
        "registrations": registrations
    }
    with open(REG_FILE, 'w') as f:
        json.dump(data, f, indent=4)

def load_row_tracker():
    """Loads the next row number for each sheet."""
    if not os.path.exists(ROW_TRACKER_FILE):
        return {}
    with open(ROW_TRACKER_FILE, 'r') as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}

def save_row_tracker(tracker):
    """Saves the next row number for each sheet."""
    with open(ROW_TRACKER_FILE, 'w') as f:
        json.dump(tracker, f, indent=4)

# Initialize current season from the saved file
_initial_data = load_data()
CURRENT_SEASON = _initial_data.get("season", "FALL2026")

# Cache workbook in memory to avoid reloading on every snipe
_workbook_cache = None
_workbook_cache_season = None

def get_workbook():
    """Returns cached workbook, reloading only if season changed."""
    global _workbook_cache, _workbook_cache_season

    if _workbook_cache is not None and _workbook_cache_season == CURRENT_SEASON:
        return _workbook_cache

    if not os.path.exists(EXCEL_FILE):
        _workbook_cache = openpyxl.Workbook()
    else:
        _workbook_cache = openpyxl.load_workbook(EXCEL_FILE)

    _workbook_cache_season = CURRENT_SEASON
    return _workbook_cache

def get_display_name(user_id, default_name):
    """Returns registered name or discord username."""
    data = load_data()
    regs = data.get("registrations", {})
    return regs.get(str(user_id), default_name)

async def download_attachment(attachment: discord.Attachment) -> str:
    """Downloads attachment and saves it locally. Returns the local path."""
    max_size = 10 * 1024 * 1024  # 10 MB limit
    if attachment.size > max_size:
        raise ValueError(f"File too large: {attachment.size / 1024 / 1024:.1f}MB (max 10MB)")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{attachment.filename}"
    filepath = os.path.join(PROOFS_DIR, filename)

    await attachment.save(filepath)
    return filepath

# --- EXCEL LOGIC ---

async def save_to_excel(sniper_name, sniper_id, number, snipee_name, snipee_id, proof_path):
    """Saves snipe data to the specific season tab in Excel."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    workbook = get_workbook()

    # Check if season tab exists, else create it
    if CURRENT_SEASON in workbook.sheetnames:
        sheet = workbook[CURRENT_SEASON]
    else:
        sheet = workbook.create_sheet(CURRENT_SEASON)
        sheet.append(["Sniper", "Points", "Snipee", "Timestamp", "Proof Link", "Sniper ID", "Snipee ID"])

    tracker = load_row_tracker()
    next_row = tracker.get(CURRENT_SEASON, 2)

    # Safeguard: ensure next_row is at least 2
    if next_row < 2:
        next_row = 2

    sheet.cell(row=next_row, column=1).value = sniper_name
    sheet.cell(row=next_row, column=2).value = number
    sheet.cell(row=next_row, column=3).value = snipee_name
    sheet.cell(row=next_row, column=4).value = timestamp
    sheet.cell(row=next_row, column=5).value = proof_path
    sheet.cell(row=next_row, column=6).value = str(sniper_id)
    sheet.cell(row=next_row, column=7).value = str(snipee_id)

    await save_workbook_async(workbook)

    tracker[CURRENT_SEASON] = next_row + 1
    save_row_tracker(tracker)

# --- BOT SETUP ---

intents = discord.Intents.default()
intents.message_content = True
intents.members = True
bot = commands.Bot(command_prefix="!", intents=intents)

async def backup_task():
    """Run scheduled backups indefinitely."""
    while True:
        schedule.run_pending()
        await asyncio.sleep(60)

_backup_scheduled = False

def schedule_backup():
    """Schedule backup every hour (prevents duplicate schedules)."""
    global _backup_scheduled
    if _backup_scheduled:
        return
    schedule.every().hour.do(lambda: asyncio.create_task(async_backup()))
    _backup_scheduled = True
    print(f"[BACKUP] Scheduled hourly backups starting at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

async def async_backup():
    """Async wrapper for backup. Merges CSV→Excel before uploading."""
    try:
        loop = asyncio.get_event_loop()
        print(f"[BACKUP] Starting backup at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        # Merge CSV into Excel (current season only)
        await loop.run_in_executor(
            executor,
            merge_csv_to_excel,
            CURRENT_SEASON,
            EXCEL_FILE
        )
        # Upload to Google Drive
        await loop.run_in_executor(executor, upload_backup, EXCEL_FILE)
        print(f"[BACKUP] Hourly backup completed successfully at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    except Exception as e:
        print(f"[BACKUP] Error during backup: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()

_backup_task_running = False

@bot.event
async def on_ready():
    global _backup_task_running
    print(f'Logged in as {bot.user}!')
    try:
        synced = await bot.tree.sync()
        print(f"Synced {len(synced)} command(s).")
    except Exception as e:
        print(e)

    # Schedule backup and start task (only once)
    schedule_backup()
    if not _backup_task_running:
        _backup_task_running = True
        bot.loop.create_task(backup_task())

# --- ADMIN COMMANDS ---

@bot.tree.command(name="change_season", description="Update the active Excel tab name (Owner Only)")
@app_commands.describe(new_season="The new season name (e.g., FALL2026)")
async def change_season(interaction: discord.Interaction, new_season: str):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    global CURRENT_SEASON
    old_season = CURRENT_SEASON
    CURRENT_SEASON = new_season.upper()
    
    # Persist the change
    data = load_data()
    save_data(CURRENT_SEASON, data.get("registrations", {}))

    await interaction.response.send_message(
        f"✅ **Season Updated!**\nOld: `{old_season}`\nNew: `{CURRENT_SEASON}`\n"
        f"Data will now be logged in the `{CURRENT_SEASON}` tab.", 
        ephemeral=True
    )

@bot.tree.command(name="register", description="Assign a custom name to a Discord user (Owner Only)")
@app_commands.describe(user="The Discord user", name="The real name to use in Excel")
async def register(interaction: discord.Interaction, user: discord.User, name: str):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    data = load_data()
    regs = data.get("registrations", {})
    regs[str(user.id)] = name
    save_data(CURRENT_SEASON, regs)
    
    await interaction.response.send_message(f"✅ Registered **{user.name}** as **{name}**.", ephemeral=True)

@bot.tree.command(name="deregister", description="Remove a custom name registration (Owner Only)")
@app_commands.describe(name="The registered name to remove")
async def deregister(interaction: discord.Interaction, name: str):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    data = load_data()
    regs = data.get("registrations", {})
    
    user_id_to_remove = next((uid for uid, n in regs.items() if n == name), None)
    
    if user_id_to_remove:
        del regs[user_id_to_remove]
        save_data(CURRENT_SEASON, regs)
        await interaction.response.send_message(f"🗑️ Removed registration for **{name}**.", ephemeral=True)
    else:
        await interaction.response.send_message(f"❌ No registration found for **{name}**.", ephemeral=True)

@deregister.autocomplete('name')
async def deregister_autocomplete(interaction: discord.Interaction, current: str):
    data = load_data()
    names = list(data.get("registrations", {}).values())
    return [
        app_commands.Choice(name=n, value=n)
        for n in names if current.lower() in n.lower()
    ][:25]

@bot.tree.command(name="list_users", description="Show all registered users (Owner Only)")
async def list_users(interaction: discord.Interaction):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    data = load_data()
    regs = data.get("registrations", {})

    if not regs:
        await interaction.response.send_message("No registered users yet.", ephemeral=True)
        return

    user_list = "\n".join([f"• {name} (ID: {uid})" for uid, name in regs.items()])
    await interaction.response.send_message(
        f"**Registered Users ({len(regs)}):**\n```\n{user_list}\n```",
        ephemeral=True
    )

@bot.tree.command(name="list_snipes", description="Show recent snipes")
@app_commands.describe(count="Number of recent snipes to show (default: 10)")
async def list_snipes(interaction: discord.Interaction, count: int = 10):
    await interaction.response.defer(ephemeral=True)

    try:
        # Read from CSV instead of Excel
        snipes = get_snipes(CURRENT_SEASON)

        if not snipes:
            await interaction.followup.send(f"No snipes recorded in `{CURRENT_SEASON}` yet.", ephemeral=True)
            return

        # Get the last N snipes (most recent first)
        snipes_to_show = snipes[-count:]
        snipes_to_show.reverse()

        snipe_lines = []
        for idx, snipe in enumerate(snipes_to_show, 1):
            sniper = snipe['Sniper']
            points = snipe['Points']
            snipee = snipe['Snipee']
            timestamp = snipe['Timestamp']
            snipe_lines.append(
                f"**{idx}.** {sniper} sniped {snipee} | {points}pts | {timestamp}"
            )

        message = "\n".join(snipe_lines)
        await interaction.followup.send(
            f"**Recent Snipes (Last {len(snipe_lines)}):**\n{message}",
            ephemeral=True
        )
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)

@bot.tree.command(name="edit_snipe", description="Edit a recent snipe (Owner Only)")
@app_commands.describe(
    row="Row number from /list_snipes (use this to identify the snipe)",
    field="Which field to edit: sniper, snipee, points, or timestamp",
    value="The new value"
)
@app_commands.choices(field=[
    app_commands.Choice(name="sniper", value="sniper"),
    app_commands.Choice(name="snipee", value="snipee"),
    app_commands.Choice(name="points", value="points"),
    app_commands.Choice(name="timestamp", value="timestamp")
])
async def edit_snipe(interaction: discord.Interaction, row: int, field: str, value: str):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    await interaction.response.defer(ephemeral=True)

    try:
        # Read from CSV instead of Excel
        csv_snipes = get_snipes(CURRENT_SEASON)

        if not csv_snipes:
            await interaction.followup.send("No snipes to edit.", ephemeral=True)
            return

        if row < 1 or row > len(csv_snipes):
            await interaction.followup.send(f"❌ Invalid row number. Use a number from 1 to {len(csv_snipes)}.", ephemeral=True)
            return

        # Map field names
        field_map = {
            "sniper": "Sniper",
            "points": "Points",
            "snipee": "Snipee",
            "timestamp": "Timestamp"
        }

        if field not in field_map:
            await interaction.followup.send("❌ Invalid field. Use: sniper, snipee, points, or timestamp.", ephemeral=True)
            return

        # Get the snipe to edit (accounting for reverse order: most recent first)
        snipe_to_edit = csv_snipes[-row]
        sniper_match = snipe_to_edit['Sniper']
        snipee_match = snipe_to_edit['Snipee']
        points_match = snipe_to_edit['Points']

        # Validate points is a number if editing points
        if field == "points":
            try:
                value = int(value)
            except ValueError:
                await interaction.followup.send("❌ Points must be a number.", ephemeral=True)
                return

        old_value = snipe_to_edit[field_map[field]]

        # Edit in CSV
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            executor,
            csv_edit_snipe,
            CURRENT_SEASON,
            sniper_match,
            snipee_match,
            points_match,
            field_map[field],
            str(value)
        )

        await interaction.followup.send(
            f"✅ **Updated:**\n"
            f"Field: `{field}`\n"
            f"Old value: `{old_value}`\n"
            f"New value: `{value}`",
            ephemeral=True
        )
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)

@bot.tree.command(name="download_sheet", description="Download the current Excel sheet (Owner Only)")
async def download_sheet(interaction: discord.Interaction):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    await interaction.response.defer(ephemeral=True)

    try:
        # Merge CSV into Excel (current season only)
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            executor,
            merge_csv_to_excel,
            CURRENT_SEASON,
            EXCEL_FILE
        )

        # Send the file as an attachment
        excel_file = discord.File(EXCEL_FILE, filename=f"SNIPESSTATS_{CURRENT_SEASON}.xlsx")
        await interaction.followup.send(
            f"📊 **Current Sheet ({CURRENT_SEASON}):**",
            file=excel_file,
            ephemeral=True
        )
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)

@bot.tree.command(name="backup_now", description="Manually trigger a backup to Google Drive (Owner Only)")
async def backup_now(interaction: discord.Interaction):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    await interaction.response.defer(ephemeral=True)

    try:
        await interaction.followup.send("⏳ Backup in progress...", ephemeral=True)
        await async_backup()
        await interaction.followup.send("✅ Backup completed successfully!", ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"❌ Backup failed: {str(e)}", ephemeral=True)

@bot.tree.command(name="backup_status", description="Check backup schedule status (Owner Only)")
async def backup_status(interaction: discord.Interaction):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    await interaction.response.defer(ephemeral=True)

    try:
        jobs = schedule.jobs
        if not jobs:
            await interaction.followup.send(
                "⚠️ **No backups scheduled!**\nRestart the bot to enable automatic backups.",
                ephemeral=True
            )
            return

        job_info = f"**Backup Schedule Status:**\n"
        job_info += f"Scheduled jobs: {len(jobs)}\n"
        for job in jobs:
            job_info += f"• Next run: {job.next_run.strftime('%Y-%m-%d %H:%M:%S')}\n"

        await interaction.followup.send(job_info, ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)

@bot.tree.command(name="delete_snipe", description="Delete a snipe from the logs (Owner Only)")
@app_commands.describe(row="Row number from /list_snipes")
async def delete_snipe(interaction: discord.Interaction, row: int):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    await interaction.response.defer(ephemeral=True)

    try:
        workbook = get_workbook()

        if CURRENT_SEASON not in workbook.sheetnames:
            await interaction.followup.send("No snipes to delete.", ephemeral=True)
            return

        sheet = workbook[CURRENT_SEASON]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= 1:
            await interaction.followup.send("No snipes to delete.", ephemeral=True)
            return

        # Read from CSV instead of Excel
        csv_snipes = get_snipes(CURRENT_SEASON)

        if not csv_snipes:
            await interaction.followup.send("No snipes to delete.", ephemeral=True)
            return

        if row < 1 or row > len(csv_snipes):
            await interaction.followup.send(f"❌ Invalid row number. Use a number from 1 to {len(csv_snipes)}.", ephemeral=True)
            return

        # Get the snipe to delete (accounting for reverse order: most recent first)
        snipe_to_delete = csv_snipes[-row]
        sniper = snipe_to_delete['Sniper']
        snipee = snipe_to_delete['Snipee']
        points = snipe_to_delete['Points']

        # Delete from CSV
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            executor,
            csv_delete_snipe,
            CURRENT_SEASON,
            sniper,
            snipee,
            points
        )

        await interaction.followup.send(
            f"🗑️ **Deleted Snipe:**\n"
            f"{sniper} → {snipee} ({points}pts)",
            ephemeral=True
        )
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)

# --- MAIN GAME COMMAND ---

@bot.tree.command(name="log_snipe", description="Manually log a snipe from a Discord message (Owner Only)")
@app_commands.describe(
    message_link="Discord message link (https://discord.com/channels/...)",
    sniper="Name of the sniper",
    snipee="Name of the snipee",
    points="Points value (1, 2, or 5)"
)
@app_commands.choices(points=[
    app_commands.Choice(name="1", value=1),
    app_commands.Choice(name="2", value=2),
    app_commands.Choice(name="Alumni Snipe", value=5)
])
async def log_snipe(interaction: discord.Interaction, message_link: str, sniper: str, snipee: str, points: int):
    if interaction.user.id != OWNER_ID:
        await interaction.response.send_message("You don't have permission for this.", ephemeral=True)
        return

    await interaction.response.defer(ephemeral=True)

    try:
        # Parse message link: https://discord.com/channels/GUILD_ID/CHANNEL_ID/MESSAGE_ID
        try:
            parts = message_link.strip('/').split('/')
            # Expected format after split: ['https:', '', 'discord.com', 'channels', GUILD_ID, CHANNEL_ID, MESSAGE_ID]
            if 'channels' not in parts:
                raise ValueError("Invalid format")

            # Get the last 3 parts as guild_id, channel_id, message_id
            message_id = int(parts[-1])
            channel_id = int(parts[-2])
            guild_id = int(parts[-3])
        except (ValueError, IndexError):
            await interaction.followup.send("❌ Invalid message link format. Use: `https://discord.com/channels/GUILD/CHANNEL/MESSAGE`", ephemeral=True)
            return

        # Fetch the message
        try:
            channel = await interaction.client.fetch_channel(channel_id)
            message = await channel.fetch_message(message_id)
        except discord.NotFound:
            await interaction.followup.send("❌ Could not find message or channel.", ephemeral=True)
            return
        except discord.Forbidden:
            await interaction.followup.send("❌ Bot does not have permission to access that channel.", ephemeral=True)
            return
        if not message.attachments:
            await interaction.followup.send("❌ Message has no attachments/proof image.", ephemeral=True)
            return

        # Download the attachment
        proof_attachment = message.attachments[0]
        proof_path = await download_attachment(proof_attachment)

        # Append to CSV
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            executor,
            append_snipe,
            CURRENT_SEASON,
            sniper,
            points,
            snipee,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            proof_path,
            interaction.user.id,
            0  # snipee_id unknown for manual logs
        )

        await interaction.followup.send(
            f"✅ **Logged Snipe:**\n"
            f"{sniper} → {snipee} ({points}pts)\n"
            f"Proof: `{os.path.basename(proof_path)}`",
            ephemeral=True
        )
    except ValueError:
        await interaction.followup.send("❌ Invalid message link or points value.", ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"❌ Error: {str(e)}", ephemeral=True)

@bot.tree.command(name="snipe", description="Add a Snipe to the Excel Sheet")
@app_commands.describe(number="Points value", user="Who did you snipe? (Leave blank for Alumni)", proof="Photo proof")
@app_commands.choices(number=[
    app_commands.Choice(name="1", value=1),
    app_commands.Choice(name="2", value=2),
    app_commands.Choice(name="Alumni Snipe", value=5)
])
async def snipe(interaction: discord.Interaction, number: int, proof: discord.Attachment, user: discord.User = None):
    # Check if command is being run in the correct channel (BEFORE defer)
    if interaction.channel.name != "ssnipes":
        await interaction.response.send_message(f"❌ Snipes must be recorded in #ssnipes channel.", ephemeral=True)
        return

    sniper_display = get_display_name(interaction.user.id, interaction.user.name)
    sniper_id = interaction.user.id

    # Handle Alumni logic vs Standard Snipe
    if user is None:
        if number == 5:
            snipee_display = "Alumni"
            snipee_id = "0000"
            display_message = f"**{sniper_display} got an Alumni Snipe for 5 points!**"
        else:
            await interaction.response.send_message("❌ You must select a user unless it is an Alumni Snipe (5 pts).", ephemeral=True)
            return
    else:
        data = load_data()
        regs = data.get("registrations", {})

        # Check if the snipee is registered (BEFORE defer)
        if str(user.id) not in regs:
            await interaction.response.send_message(
                f"❌ **{user.name}** is not registered. Ask an admin to register them first using `/register`.",
                ephemeral=True
            )
            return

        snipee_display = get_display_name(user.id, user.name)
        snipee_id = user.id
        display_message = f"**<@{snipee_id}> ({snipee_display}) got shot by {sniper_display} for {number} points**"

    # All validation passed, now defer
    await interaction.response.defer()

    print(f"[SNIPE] About to download attachment: {proof.filename} ({proof.size} bytes)")
    try:
        proof_path = await download_attachment(proof)

        # Save to CSV instead of Excel (fast!)
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            executor,
            append_snipe,
            CURRENT_SEASON,
            sniper_display,
            number,
            snipee_display,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            proof_path,
            sniper_id,
            snipee_id
        )

        proof_file = discord.File(proof_path, filename=os.path.basename(proof_path))
        await interaction.followup.send(f"{display_message}", file=proof_file)
    except ValueError as e:
        await interaction.followup.send(f"❌ **File Error:** {str(e)}")
    except PermissionError:
        await interaction.followup.send(f"⚠️ <@{OWNER_ID}> **CLOSE THE EXCEL SHEET**")
    except Exception as e:
        await interaction.followup.send(f"❌ **TECHNICAL ERROR:** `{str(e)}`")
        print(f"Error in snipe command: {e}")

bot.run(TOKEN)